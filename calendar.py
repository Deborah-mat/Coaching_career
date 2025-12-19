import streamlit as st
import pandas as pd
from pathlib import Path
import plotly.graph_objects as go
import io
from openpyxl import Workbook
from openpyxl.styles import Font
import base64
import plotly.io as pio

# =============================
# PATHS
# =============================
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

# =============================
# GREEK TIME ARCHETYPES
# =============================
GREEK_ARCHETYPES = {
    "CHR – Chronos": {
        "description": "Sequential, measurable clock time",
        "color": "#1f77b4",
        "symbol": "⏱️"
    },
    "KAI – Kairos": {
        "description": "Perfect timing, opportune moments",
        "color": "#ff7f0e",
        "symbol": "🎯"
    },
    "AIO – Aion": {
        "description": "Timeless, eternal perspective",
        "color": "#2ca02c",
        "symbol": "♾️"
    },
    "ANA – Ananke": {
        "description": "Necessity, urgent obligations",
        "color": "#d62728",
        "symbol": "⚡"
    },
    "HEL – Helios": {
        "description": "Daylight, active productive hours",
        "color": "#ffd700",
        "symbol": "☀️"
    },
    "NYX – Nyx": {
        "description": "Night-time, rest and shadow work",
        "color": "#4b0082",
        "symbol": "🌙"
    },
    "EOS – Eos": {
        "description": "Dawn, fresh starts and new beginnings",
        "color": "#ff69b4",
        "symbol": "🌅"
    },
    "HOR – Horae": {
        "description": "Natural rhythms, seasons and cycles",
        "color": "#32cd32",
        "symbol": "🌱"
    },
    "MNE – Mnemosyne": {
        "description": "Memory work, reflection and learning",
        "color": "#8a2be2",
        "symbol": "💭"
    },
    "HYP – Hypnos": {
        "description": "Sleep, deep rest and recovery",
        "color": "#4682b4",
        "symbol": "😴"
    },
    "ONE – Oneiros": {
        "description": "Dreams, imagination and creative play",
        "color": "#da70d6",
        "symbol": "✨"
    },
    "Movement/Commute": {
        "description": "Travel and transportation",
        "color": "#7f7f7f",
        "symbol": "🚌"
    },
    "(empty)": {
        "description": "Free/unscheduled time",
        "color": "#f0f0f0",
        "symbol": "⭕"
    }
}

# =============================
# STREAMLIT CONFIG
# =============================
st.set_page_config(page_title="📅 Personalised Schedule", layout="wide")
st.title("📅 Personalised Schedule – Greek Time Archetypes Calendar")
st.caption("Hover over time slots for details.")

# =============================
# FONCTIONS
# =============================
def sort_days(days_list):
    """Sort days of the week in correct order"""
    day_order = {
        'monday': 1, 'mon': 1, 'lundi': 1, 'lun': 1,
        'tuesday': 2, 'tue': 2, 'mardi': 2, 'mar': 2,
        'wednesday': 3, 'wed': 3, 'mercredi': 3, 'mer': 3,
        'thursday': 4, 'thu': 4, 'jeudi': 4, 'jeu': 4,
        'friday': 5, 'fri': 5, 'vendredi': 5, 'ven': 5,
        'saturday': 6, 'sat': 6, 'samedi': 6, 'sam': 6,
        'sunday': 7, 'sun': 7, 'dimanche': 7, 'dim': 7
    }
    
    # Filtrer les valeurs vides
    filtered_days = []
    for d in days_list:
        if pd.notna(d):
            d_str = str(d).strip()
            if d_str and d_str.lower() != "nan":
                filtered_days.append(d)
    
    # Trier en fonction de l'ordre des jours
    def get_day_order_key(day_str):
        if pd.isna(day_str) or not str(day_str).strip():
            return 99
        day_str_clean = str(day_str).strip().lower()
        parts = day_str_clean.split()
        if not parts:
            return 99
        return day_order.get(parts[0], 99)
    
    sorted_days = sorted(filtered_days, key=get_day_order_key)
    
    return sorted_days

def parse_time_minutes(time_str):
    """Convert HH:MM to minutes"""
    try:
        if pd.isna(time_str):
            return None
        time_str = str(time_str).strip()
        if ":" in time_str:
            parts = time_str.split(":")
            if len(parts) >= 2:
                return int(parts[0]) * 60 + int(parts[1])
        return None
    except:
        return None

def normalize_dataframe(df, file_name):
    """Normalize column names for consistency"""
    df = df.copy()
    
    # Create mapping dictionary based on actual column names
    column_mapping = {}
    
    for col in df.columns:
        col_str = str(col).strip()
        col_lower = col_str.lower()
        
        # Map to standard names
        if 'date' in col_lower or col_lower == 'day':
            column_mapping[col] = 'day'
        elif 'activity' in col_lower or 'title' in col_lower:
            column_mapping[col] = 'title'
        elif 'start' in col_lower:
            column_mapping[col] = 'start'
        elif 'end' in col_lower:
            column_mapping[col] = 'end'
        elif 'archetype' in col_lower:
            column_mapping[col] = 'archetype'
        elif 'note' in col_lower or 'detail' in col_lower or 'exam' in col_lower:
            column_mapping[col] = 'notes'
    
    # Apply mapping
    df = df.rename(columns=column_mapping)
    
    # Ensure all required columns exist
    required = ['day', 'title', 'start', 'end']
    for col in required:
        if col not in df.columns:
            df[col] = ""
    
    if 'archetype' not in df.columns:
        df['archetype'] = "CHR – Chronos"
    if 'notes' not in df.columns:
        df['notes'] = ""
    
    # Clean data
    for col in df.columns:
        df[col] = df[col].fillna("").astype(str).str.strip()
    
    # Debug info
    st.sidebar.info(f"📊 {file_name}")
    st.sidebar.write(f"Columns: {list(df.columns)}")
    
    return df

def create_calendar_heatmap(df, days, week_key, show_text=True):
    """Create calendar heatmap"""
    if df.empty:
        st.warning("No data to display")
        return None
    
    # Time slots from 6:00 to 23:00
    slot_minutes = 30
    start_hour = 6
    end_hour = 23
    
    slots = list(range(start_hour * 60, end_hour * 60, slot_minutes))
    y_labels = [f"{m//60:02d}:{m%60:02d}" for m in slots]
    
    # Get unique archetypes
    all_archetypes = sorted(set(df["archetype"].unique().tolist() + ["(empty)"]))
    arch_to_idx = {a: i for i, a in enumerate(all_archetypes)}
    
    # Prepare data for heatmap
    z, text, hover = [], [], []
    
    for slot in slots:
        row_z, row_t, row_h = [], [], []
        
        for day in days:
            # Find events for this day and time slot
            day_str = str(day).strip()
            day_events = df[df["day"].str.strip().str.lower() == day_str.lower()]
            matching_events = []
            
            for _, event in day_events.iterrows():
                start_min = parse_time_minutes(event["start"])
                end_min = parse_time_minutes(event["end"])
                
                if start_min and end_min and start_min <= slot < end_min:
                    matching_events.append(event)
            
            if not matching_events:
                # Empty slot
                row_z.append(arch_to_idx["(empty)"])
                row_t.append("")
                row_h.append(f"<b>{day}</b><br>Free time<br><i>No scheduled activities</i>")
            else:
                # Use first matching event
                event = matching_events[0]
                archetype = event["archetype"]
                row_z.append(arch_to_idx[archetype])
                
                # Display text
                arch_info = GREEK_ARCHETYPES.get(archetype, {"symbol": ""})
                symbol = arch_info.get("symbol", "")
                display_text = f"{symbol} {event['title']}" if symbol else event['title']
                row_t.append(display_text[:30])
                
                # Hover text
                hover_text = f"""
                <b>{day}</b><br>
                <b>⏰ {event['start']} - {event['end']}</b><br>
                📝 <b>{event['title']}</b><br>
                🏛️ <i>{archetype}</i><br>
                📋 {GREEK_ARCHETYPES.get(archetype, {}).get('description', '')}<br>
                """
                
                if pd.notna(event.get('notes')) and str(event['notes']).strip():
                    hover_text += f"📎 <i>Notes: {event['notes']}</i><br>"
                
                row_h.append(hover_text)
        
        z.append(row_z)
        text.append(row_t)
        hover.append(row_h)
    
    # Create colorscale
    colorscale = []
    for i, arch in enumerate(all_archetypes):
        color = GREEK_ARCHETYPES.get(arch, {"color": "#cccccc"})["color"]
        colorscale.append([i/len(all_archetypes), color])
    
    # Create figure
    fig = go.Figure(go.Heatmap(
        z=z,
        x=days,
        y=y_labels,
        text=text if show_text else None,
        hovertext=hover,
        hoverinfo="text",
        colorscale=colorscale,
        showscale=False,
        texttemplate="%{text}" if show_text else None,
        textfont=dict(size=10, color="black"),
        hovertemplate="%{hovertext}<extra></extra>"
    ))
    
    fig.update_layout(
        height=800,
        margin=dict(l=80, r=30, t=80, b=30),
        xaxis_title="<b>Days</b>",
        yaxis_title="<b>Time</b>",
        title=dict(
            text=f"📅 {week_key.capitalize()} Week Schedule",
            font=dict(size=18, color='darkblue'),
            x=0.5
        )
    )
    
    fig.update_yaxes(autorange="reversed", showgrid=True)
    
    return fig

def export_calendar_as_png(df, days, week_key):
    """Export as PNG"""
    fig = create_calendar_heatmap(df, days, week_key, show_text=True)
    
    if fig is None:
        return None
    
    # Adjust for export
    fig.update_layout(
        height=1200,
        width=1600,
        margin=dict(l=100, r=100, t=150, b=150),
        title=dict(
            text=f"📅 {week_key.capitalize()} Week Schedule - EXPORT",
            font=dict(size=24)
        )
    )
    
    # Convert to PNG
    try:
        img_bytes = pio.to_image(fig, format="png", width=1600, height=1200, scale=2, engine="kaleido")
        return io.BytesIO(img_bytes)
    except Exception as e:
        st.error(f"Error generating PNG: {e}")
        return None

def create_excel_export(df, week_type, days):
    """Create Excel export"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{week_type} Schedule"
    
    # Title
    ws['A1'] = f"📋 {week_type.capitalize()} Week Schedule"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Headers
    headers = ["Day", "Time", "Activity", "Archetype", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Data
    row = 4
    for day in days:
        day_str = str(day).strip()
        day_events = df[df["day"].str.strip().str.lower() == day_str.lower()]
        
        if day_events.empty:
            ws.cell(row=row, column=1, value=day)
            ws.cell(row=row, column=3, value="Free time")
            row += 1
        else:
            for _, event in day_events.iterrows():
                ws.cell(row=row, column=1, value=day)
                ws.cell(row=row, column=2, value=f"{event['start']}-{event['end']}")
                ws.cell(row=row, column=3, value=event['title'])
                ws.cell(row=row, column=4, value=event['archetype'])
                if pd.notna(event.get('notes')):
                    ws.cell(row=row, column=5, value=str(event['notes']))
                row += 1
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# =============================
# MAIN APP
# =============================
def main():
    # Display legend in sidebar
    st.sidebar.header("🏛️ Greek Time Archetypes")
    for archetype, info in GREEK_ARCHETYPES.items():
        if archetype != "(empty)":
            color = info["color"]
            symbol = info.get("symbol", "")
            st.sidebar.markdown(
                f'<div style="background-color:{color}20; padding:8px; margin:5px 0; border-left:4px solid {color}">'
                f'<strong>{symbol} {archetype}</strong><br>'
                f'<small>{info["description"]}</small>'
                f'</div>',
                unsafe_allow_html=True
            )
    
    st.sidebar.markdown("---")
    st.sidebar.header("📂 Load Schedule Data")
    
    # Option 1: Load from files in data directory
    busy_path = DATA_DIR / "busy_week.xlsx"
    quiet_path = DATA_DIR / "quiet_week.xlsx"
    
    busy_df = pd.DataFrame()
    quiet_df = pd.DataFrame()
    
    if busy_path.exists():
        try:
            raw_busy = pd.read_excel(busy_path)
            busy_df = normalize_dataframe(raw_busy, "busy_week.xlsx")
            st.sidebar.success(f"✅ Loaded busy week")
        except Exception as e:
            st.sidebar.error(f"❌ Error loading busy week: {e}")
    
    if quiet_path.exists():
        try:
            raw_quiet = pd.read_excel(quiet_path)
            quiet_df = normalize_dataframe(raw_quiet, "quiet_week.xlsx")
            st.sidebar.success(f"✅ Loaded quiet week: {len(quiet_df)} events")
        except Exception as e:
            st.sidebar.error(f"❌ Error loading quiet week: {e}")
    
    # Option 2: Upload files directly
    uploaded_files = st.sidebar.file_uploader(
        "Or upload Excel files", 
        type=["xlsx", "xls"],
        accept_multiple_files=True
    )
    
    if uploaded_files:
        for uploaded_file in uploaded_files:
            try:
                df = pd.read_excel(uploaded_file)
                normalized_df = normalize_dataframe(df, uploaded_file.name)
                
                if "busy" in uploaded_file.name.lower():
                    busy_df = normalized_df
                    st.sidebar.success(f"✅ Loaded {uploaded_file.name} as busy week")
                elif "quiet" in uploaded_file.name.lower():
                    quiet_df = normalized_df
                    st.sidebar.success(f"✅ Loaded {uploaded_file.name} as quiet week")
                else:
                    # Default to busy if name unclear
                    busy_df = normalized_df
                    st.sidebar.info(f"📄 Loaded {uploaded_file.name}")
            except Exception as e:
                st.sidebar.error(f"❌ Error loading {uploaded_file.name}: {e}")
    
    # Create tabs for busy/quiet weeks
    if not busy_df.empty or not quiet_df.empty:
        tab1, tab2 = st.tabs(["🔥 Busy Week", "🌿 Quiet Week"])
        
        with tab1:
            if not busy_df.empty:
                # Utiliser la fonction sort_days pour trier correctement
                days = sort_days(busy_df["day"].unique())
                
                if days:
                    st.subheader(f"🔥 Busy Week Schedule ({len(busy_df)} events)")
                    st.write(f"**Days:** {', '.join(str(d) for d in days)}")
                    
                    # Show raw data for debugging
                    with st.expander("🔍 View raw data"):
                        st.dataframe(busy_df)
                    
                    show_text = st.checkbox("Show text in calendar", value=True, key="busy_text")
                    
                    fig = create_calendar_heatmap(busy_df, days, "busy", show_text)
                    if fig:
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Export buttons - CORRECTION ICI
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            # Bouton PNG avec download_button au lieu de markdown
                            if st.button("📥 Download PNG", key="png_busy"):
                                png_buffer = export_calendar_as_png(busy_df, days, "busy")
                                if png_buffer:
                                    st.download_button(
                                        label="📥 Click to download PNG",
                                        data=png_buffer,
                                        file_name="busy_week_schedule.png",
                                        mime="image/png",
                                        key="download_png_busy"
                                    )
                        
                        with col2:
                            # Bouton Excel
                            if st.button("📊 Download Excel", key="excel_busy"):
                                excel_buffer = create_excel_export(busy_df, "busy", days)
                                if excel_buffer:
                                    st.download_button(
                                        label="📥 Click to download Excel",
                                        data=excel_buffer,
                                        file_name="busy_week_schedule.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key="download_excel_busy"
                                    )
                else:
                    st.warning("No valid days found in the data. Please check your 'day' column.")
            else:
                st.info("No busy week data available. Upload a file or place busy_week.xlsx in the /data folder.")
        
        with tab2:
            if not quiet_df.empty:
                # Utiliser la fonction sort_days pour trier correctement
                days = sort_days(quiet_df["day"].unique())
                
                if days:
                    st.subheader(f"🌿 Quiet Week Schedule ({len(quiet_df)} events)")
                    st.write(f"**Days:** {', '.join(str(d) for d in days)}")
                    
                    # Show raw data for debugging
                    with st.expander("🔍 View raw data"):
                        st.dataframe(quiet_df)
                    
                    show_text = st.checkbox("Show text in calendar", value=True, key="quiet_text")
                    
                    fig = create_calendar_heatmap(quiet_df, days, "quiet", show_text)
                    if fig:
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Export buttons - CORRECTION ICI
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            # Bouton PNG avec download_button au lieu de markdown
                            if st.button("📥 Download PNG", key="png_quiet"):
                                png_buffer = export_calendar_as_png(quiet_df, days, "quiet")
                                if png_buffer:
                                    st.download_button(
                                        label="📥 Click to download PNG",
                                        data=png_buffer,
                                        file_name="quiet_week_schedule.png",
                                        mime="image/png",
                                        key="download_png_quiet"
                                    )
                        
                        with col2:
                            # Bouton Excel
                            if st.button("📊 Download Excel", key="excel_quiet"):
                                excel_buffer = create_excel_export(quiet_df, "quiet", days)
                                if excel_buffer:
                                    st.download_button(
                                        label="📥 Click to download Excel",
                                        data=excel_buffer,
                                        file_name="quiet_week_schedule.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key="download_excel_quiet"
                                    )
                else:
                    st.warning("No valid days found in the data. Please check your 'day' column.")
            else:
                st.info("No quiet week data available. Upload a file or place quiet_week.xlsx in the /data folder.")
    
    else:
        st.error("""
        ❌ **No schedule data found!**
        
        **Please either:**
        1. Place Excel files in the `/data` folder:
           - `busy_week.xlsx`
           - `quiet_week.xlsx`
        
        2. **OR** upload files directly using the uploader in the sidebar.
        
        **Your files should look like:**
        
        **busy_week.xlsx:**
        ```
        | Date      | Activity          | Start Time | End Time | Archetype        | Notes (Exam Duration) |
        | Monday    | Commute to school | 08:00      | 08:45    | Movement/Commute | Bus ride              |
        ```
        
        **quiet_week.xlsx:**
        ```
        | Day       | Activity           | Start | End   | Archetype     | Details        |
        | Monday    | Morning meditation | 08:00 | 09:00 | EOS – Eos     | Daily practice |
        ```
        """)

if __name__ == "__main__":
    main()